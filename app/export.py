import math
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from . import calc
from .timeutil import to_pacific

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_sheet(ws, headers, rows):
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    for col_idx, header in enumerate(headers, start=1):
        width = max(12, min(40, len(str(header)) + 4))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_simple_workbook(run):
    """Ready-to-execute shipment list: only SKUs marked included. Column set
    mirrors the other distributor tool's packing-list form (SKU, Barcode,
    Name, 보관, 입수/박스, 박스수, 수량(EA), KPAC/HQ box split) for easy
    side-by-side comparison and familiarity for whoever executes the pull.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "EAST Shipment"

    headers = ["SKU", "Barcode", "Name", "보관", "입수/박스", "박스수", "수량(EA)", "Kpac(box)", "HQ(box)"]
    rows = []
    items = [it for it in run.items if it.included]
    items.sort(key=lambda it: (it.storage or "", it.sku))
    total_boxes = 0.0
    for it in items:
        qty = it.effective_qty()
        kpac_boxes, hq_boxes, hq_units = it.kpac_hq_split()
        if it.box_size:
            total_boxes += qty / it.box_size
        rows.append([
            it.sku, it.barcode, it.name, it.storage, it.box_size or "", kpac_boxes + hq_boxes, qty,
            kpac_boxes, hq_boxes,
        ])

    total_pallets = total_boxes / calc.BOXES_PER_PALLET
    aoa_extra = [
        [],
        ["Total Est. Boxes (all included SKUs)", round(total_boxes, 2)],
        [f"Total Est. Pallets ({calc.BOXES_PER_PALLET} boxes/pallet)", round(total_pallets, 2)],
    ]

    _write_sheet(ws, headers, rows)
    for row in aoa_extra:
        ws.append(row)
    return wb


def build_detailed_workbook(run):
    """Full calculation breakdown per SKU, including every intermediate
    value and the reasoning/formula that produced the final quantity.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Detailed Calculation"

    settings = run.settings()
    cap_pct = settings.get("cap_pct", 0.40)
    floor_pct = settings.get("floor_pct", 0.25)
    target_days = settings.get("target_days", 21)
    box_round_pct = settings.get("box_round_pct", 0.30)

    headers = [
        "SKU", "Barcode", "Name", "English Name", "Category", "온라인 중요도", "Status.Online",
        "HQ_Online", "LOCAL_TRANS_ONLINE", "KPAC_Online", "HQ_qty (sum)",
        "PA_ART", "PA_ONLINE", "EAST_onhand (sum)",
        "PA_1_PRE_STORAGE", "PA_2_PRE_STORAGE", "In-transit (MAX)", "In-transit ignored (MIN, stale)",
        "Total Units (HQ+EAST+Transit)", "EAST Position (onhand+transit)",
        "5d avg/day", "30d avg/day", "90d avg/day", "Window used (max)", "Total avg/day used",
        "EAST avg/day (actual EAST-only sales, matching window)",
        f"Desired Qty (top up to {target_days}d of EAST avg/day, minus EAST Position)",
        f"Min Floor Qty (max(0, {floor_pct:.0%}*Total - EAST Pos))",
        "Target Qty (max(Desired, Floor))",
        f"Cap Room ({cap_pct:.0%}*Total - EAST Pos)",
        "Raw Qty (min(Target, Cap Room, HQ_qty))",
        "Binding Constraint",
        "Box Size (입수/박스)",
        f"Box Rounding Rule (round to nearest box; round down if remainder < {box_round_pct:.0%} of box, else up; "
        f"reverts to units if that would breach the floor/cap)",
        "Final Qty (computed)",
        "Flags",
        "Included in shipment?",
        "Final Qty (user override, if any)",
        "Final Qty (effective)",
        "Est. Boxes (effective)",
        "HQ days after ship (effective)",
        "EAST days after ship (effective)",
        "KPAC Available Boxes (KPAC_Online / Box Size)",
        "KPAC Boxes to Pull",
        "HQ Boxes to Pull",
        "HQ Extra Units to Pull",
    ]

    rows = []
    total_boxes_included = 0.0
    total_kpac_boxes = 0.0
    total_hq_boxes = 0.0
    total_hq_units = 0.0
    for it in sorted(run.items, key=lambda x: x.sku):
        boxes = it.est_boxes()
        hq_days, east_days = it.days_after_ship()
        kpac_boxes, hq_boxes, hq_units = it.kpac_hq_split()
        kpac_avail = math.floor(it.kpac_online / it.box_size) if it.box_size else ""
        if it.included:
            if boxes is not None:
                total_boxes_included += boxes
            total_kpac_boxes += kpac_boxes
            total_hq_boxes += hq_boxes
            total_hq_units += hq_units
        rows.append([
            it.sku, it.barcode, it.name, it.english_name, it.category, it.importance, it.status_online,
            it.hq_online, it.local_trans_online, it.kpac_online, it.hq_qty,
            it.pa_art, it.pa_online, it.east_onhand,
            it.pa_1_pre, it.pa_2_pre, it.in_transit, it.in_transit_ignored,
            it.total_units, it.east_position,
            it.avg5, it.avg30, it.avg90, it.window_used or "n/a", it.total_avg_daily,
            round(it.east_avg_daily, 4),
            round(it.desired_qty, 2),
            round(it.min_floor_qty, 2),
            round(it.target_qty, 2),
            round(it.cap_room, 2),
            round(it.raw_qty, 2),
            it.binding_constraint,
            it.box_size or "",
            it.rounding_note,
            it.final_qty_computed,
            ", ".join(it.flags()),
            "YES" if it.included else "no",
            it.final_qty_user if it.final_qty_user is not None else "",
            it.effective_qty(),
            round(boxes, 2) if boxes is not None else "",
            round(hq_days, 2) if hq_days is not None else "",
            round(east_days, 2) if east_days is not None else "",
            kpac_avail,
            kpac_boxes if it.included else "",
            hq_boxes if it.included else "",
            round(hq_units, 2) if it.included else "",
        ])

    _write_sheet(ws, headers, rows)

    # A short methodology sheet for reference.
    notes = wb.create_sheet("Methodology")
    notes.append(["Parameter", "Value"])
    notes["A1"].font = HEADER_FONT
    notes["B1"].font = HEADER_FONT
    notes["A1"].fill = HEADER_FILL
    notes["B1"].fill = HEADER_FILL
    kimchi_settings = run.settings()
    for k, v in [
        ("EAST cap (% of total units)", f"{cap_pct:.0%}"),
        ("EAST minimum floor (% of total units)", f"{floor_pct:.0%}"),
        ("Target days of EAST coverage to top up to", target_days),
        ("Box rounding threshold (% of box size)", f"{box_round_pct:.0%}"),
        ("김치 override enabled", kimchi_settings.get("kimchi_override_enabled", True)),
        ("김치 keyword", kimchi_settings.get("kimchi_keyword", "김치")),
        ("김치 EAST minimum floor", f"{kimchi_settings.get('kimchi_floor_pct', 0.10):.0%}"),
        ("김치 EAST cap (soft)", f"{kimchi_settings.get('kimchi_cap_pct', 0.15):.0%}"),
        ("김치 EAST hard cap", f"{kimchi_settings.get('kimchi_hard_cap_pct', 0.20):.0%}"),
        ("Run created", to_pacific(run.created_at).strftime("%Y-%m-%d %H:%M PT")),
        ("Run status", run.status),
        ("Source file", run.source_filename),
        ("Total Est. Boxes (all included SKUs)", round(total_boxes_included, 2)),
        (f"Total Est. Pallets ({calc.BOXES_PER_PALLET} boxes/pallet)", round(total_boxes_included / calc.BOXES_PER_PALLET, 2)),
        ("Total KPAC Boxes to Pull (all included SKUs)", round(total_kpac_boxes, 2)),
        ("Total HQ Boxes to Pull (all included SKUs)", round(total_hq_boxes, 2)),
        ("Total HQ Extra Units to Pull (all included SKUs)", round(total_hq_units, 2)),
    ]:
        notes.append([k, v])
    notes.column_dimensions["A"].width = 40
    notes.column_dimensions["B"].width = 30

    return wb


def workbook_to_bytes(wb):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
